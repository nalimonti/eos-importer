from openpyxl import load_workbook
import MySQLdb
import MySQLdb.cursors
from datetime import datetime

db = MySQLdb.connect(host="127.0.0.1", db="eos-import", passwd="", user="root", cursorclass=MySQLdb.cursors.DictCursor)
cursor = db.cursor()
candidates = {}
duplicates = {}
excluded_ids = []

def insert_str(table, fields):
    return "INSERT INTO {} ({}) VALUES ({})".format(
        table,
        ', '.join(fields),
        ','.join(['%s' for field in fields])
    )

def dict_to_tuple(row, fields):
    val_list = []
    for field in fields:
        if field not in row or not row[field]:
            val_list.append('')
        elif field == 'DATE_OF_BIRTH':
            date = datetime.strptime(row[field], '%m/%d/%Y')
            val_list.append(date)
        else:
            val_list.append(row[field])

    return tuple(val_list)

def convert_key(key):
    if key == 'STATE':
        return 'state'
    elif key == 'ADDRESS1':
        return 'address_1'
    elif key == 'ADDRESS2':
        return 'address_2'
    elif key == 'CITY':
        return 'city'
    elif key == 'POSTAL_CODE':
        return 'zip'
    elif key == 'TELEPHONE_NUMBER':
        return 'phone'
    elif key == 'EMAIL':
        return 'email'
    elif key == 'FIRST_NAME':
        return 'first_name'
    elif key == 'LAST_NAME':
        return 'last_name'
    elif key == 'DATE_OF_BIRTH':
        return 'dob'
    elif key == 'CANDIDATE_ID':
        return 'candidate_id'

def insert(insert_str, args=[]):
    cursor.executemany(insert_str, args)
    db.commit()

def find_dupe_emails():
    wb = load_workbook('demograhpic_info_no_pii 5 17 2018.xlsx')
    ws = wb[wb.sheetnames[0]]

    seen = {}

    for i in range(2, ws.max_row):
        e = ws['G' + str(i)].value
        p_id = ws['A' + str(i)].value

        if not e:
            continue

        e = e.lower()

        if e not in seen:
            seen[e] = [p_id]
        else:
            seen[e].append(p_id)
    return {k: v for k, v in seen.items() if len(v) > 1}


def build_examinee_dict():
    global duplicates
    duplicates = find_dupe_emails()
    global excluded_ids
    excluded_ids = sorted({x for v in duplicates.values() for x in v})

    wb = load_workbook('examinee_list_pii 5 17 2018.xlsx')
    ws = wb[wb.sheetnames[0]]

    keys = [ws.cell(row=1, column=i).value for i in range(1, 7)]

    for i in range(2, ws.max_row + 1):
        p_id = ws['A' + str(i)].value

        if not p_id or p_id in excluded_ids:
            continue

        candidates[p_id] = {keys[j - 1]: ws.cell(row=i, column=j).value for j in range(2, 7)}

def add_demo_data():
    wb = load_workbook('demograhpic_info_no_pii 5 17 2018.xlsx')
    ws = wb[wb.sheetnames[0]]

    keys = [ws.cell(row=1, column=i).value for i in range(1, 15)]
    for i in range(2, ws.max_row + 1):
        p_id = ws['A' + str(i)].value

        if not p_id or p_id in excluded_ids:
            continue

        new_data = {keys[j - 1]: ws.cell(row=i, column=j).value for j in range(2, 15)}

        if p_id in candidates:
            old_data = candidates[p_id]
            candidates[p_id] = {**old_data, **new_data}
        else:
            candidates[p_id] = new_data

def add_test_data():
    wb = load_workbook('certificate_info.xlsx')
    ws = wb[wb.sheetnames[0]]

    keys = [ws.cell(row=1, column=i).value for i in range(1, 16)]

    for i in range(2, ws.max_row + 1):
        p_id = ws['A' + str(i)].value

        if not p_id or p_id not in candidates:
            continue

        candidate = candidates[p_id]
        if 'WORK_KEYS' not in candidate:
            candidate['WORK_KEYS'] = []

        test = {keys[j - 1]: ws.cell(row=i, column=j).value for j in range(2, 16)}
        candidate['WORK_KEYS'].append(test)

def insert_addresses(candidate_list):
    fields = [
        'ADDRESS1',
        'ADDRESS2',
        'CITY',
        'STATE',
        'POSTAL_CODE',
        'CANDIDATE_ID'
    ]

    address_tuples = [dict_to_tuple(row, fields) for row in candidate_list]
    insert_string = insert_str('candidate_addresses', [convert_key(key) for key in fields])
    insert(insert_string, address_tuples)

def insert_candidates():
    fields = [
        'FIRST_NAME',
        'LAST_NAME',
        'EMAIL',
        'TELEPHONE_NUMBER',
        'DATE_OF_BIRTH'
    ]

    candidate_list = [c for c in candidates.values()]

    for i, row in enumerate(candidate_list):
        row['CANDIDATE_ID'] = i + 1

    candidate_tuples = [dict_to_tuple(row, fields) for row in candidate_list]
    insert_string = insert_str('candidates', [convert_key(key) for key in fields])
    insert(insert_string, candidate_tuples)

    insert_addresses(candidate_list)



build_examinee_dict()
add_demo_data()
add_test_data()
insert_candidates()

